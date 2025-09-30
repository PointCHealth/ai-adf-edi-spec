"""
Generate Azure EDI Platform Budget Spreadsheet
Creates a detailed Excel workbook with service costs, transaction estimates, and scenarios
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_budget_spreadsheet():
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create sheets
    create_summary_sheet(wb)
    create_unit_pricing_sheet(wb)
    create_monthly_costs_sheet(wb)
    create_transaction_volumes_sheet(wb)
    create_environment_rollup_sheet(wb)
    create_sensitivity_analysis_sheet(wb)
    
    # Save workbook
    filename = f"Azure_EDI_Budget_Plan_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✓ Created: {filename}")
    return filename

def create_summary_sheet(wb):
    """Executive summary with key numbers"""
    ws = wb.create_sheet("Executive Summary", 0)
    
    # Header
    ws['A1'] = "Healthcare EDI Platform - Azure Budget Summary"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Prepared: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(size=10, italic=True)
    
    # Key metrics
    ws['A4'] = "Budget Scenarios (Production Monthly)"
    ws['A4'].font = Font(size=14, bold=True)
    
    headers = ['Scenario', 'Monthly Cost', 'Annual Cost', 'Notes']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    scenarios = [
        ['Low (Optimized)', '$439', '$5,268', 'Deferred features (no APIM, no Purview)'],
        ['Expected (Baseline)', '$1,438', '$17,256', 'Full VNet integration + API Management'],
        ['High (Peak/Contingency)', '$2,147', '$25,764', 'Peak volume + 2x Functions instances'],
    ]
    
    for row, data in enumerate(scenarios, start=7):
        for col, value in enumerate(data, start=1):
            ws.cell(row=row, column=col, value=value)
    
    # Annual budget ask
    ws['A11'] = "Year 1 Budget Request (All Environments)"
    ws['A11'].font = Font(size=14, bold=True)
    
    ws['A12'] = "Scenario"
    ws['B12'] = "Dev (25%)"
    ws['C12'] = "Test (40%)"
    ws['D12'] = "Prod"
    ws['E12'] = "Total/Month"
    ws['F12'] = "Annual"
    ws['G12'] = "+ 20% Contingency"
    
    for col in range(1, 8):
        cell = ws.cell(row=12, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    ws['A13'] = "Expected"
    ws['B13'] = "$424"
    ws['C13'] = "$603"
    ws['D13'] = "$1,438"
    ws['E13'] = "$2,465"
    ws['F13'] = "$29,580"
    ws['G13'] = "$35,500"
    
    ws['G13'].font = Font(bold=True, size=12)
    ws['G13'].fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    
    # Key assumptions
    ws['A16'] = "Key Assumptions"
    ws['A16'].font = Font(size=14, bold=True)
    
    assumptions = [
        "• VNet Integration: All services use private endpoints and managed VNet",
        "• Functions: Premium EP1 plan required for VNet integration",
        "• API Management: Standard v2 tier (50M requests/mo included)",
        "• Volume: ~5,000 files/week, ~250K claims/year",
        "• Environments: Dev, Test, Prod (3 total)",
        "• Region: Single primary region (East US assumed)",
        "• Retention: 7-year immutable storage with lifecycle tiering",
    ]
    
    for i, assumption in enumerate(assumptions, start=17):
        ws[f'A{i}'] = assumption
    
    # Top cost drivers
    ws['A26'] = "Top Cost Drivers (Expected Scenario)"
    ws['A26'].font = Font(size=14, bold=True)
    
    ws['A27'] = "Service"
    ws['B27'] = "Monthly Cost"
    ws['C27'] = "% of Total"
    
    for col in range(1, 4):
        cell = ws.cell(row=27, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    drivers = [
        ['API Management', '$700', '49%'],
        ['Purview Data Governance', '$190', '13%'],
        ['Data Factory', '$215', '15%'],
        ['Functions Premium', '$170', '12%'],
        ['Log Analytics', '$85', '6%'],
        ['Other', '$78', '5%'],
    ]
    
    for row, data in enumerate(drivers, start=28):
        for col, value in enumerate(data, start=1):
            ws.cell(row=row, column=col, value=value)
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20

def create_unit_pricing_sheet(wb):
    """Detailed unit pricing for all services"""
    ws = wb.create_sheet("Unit Pricing")
    
    ws['A1'] = "Azure Service Unit Pricing (PAYG)"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = "Retrieved: 2025-09-29 from Microsoft pricing pages"
    ws['A2'].font = Font(size=10, italic=True)
    
    headers = ['Service / Meter', 'Unit Price (USD)', 'Unit', 'Source']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    pricing_data = [
        ['Blob Storage - Hot (LRS)', '$0.018', 'per GB-month', 'Azure Blob Storage pricing'],
        ['Blob Storage - Cool (LRS)', '$0.010', 'per GB-month', 'Azure Blob Storage pricing'],
        ['Blob Storage - Cold (LRS)', '$0.0036', 'per GB-month', 'Azure Blob Storage pricing'],
        ['Blob Storage - Archive (LRS)', '$0.002', 'per GB-month', 'Azure Blob Storage pricing'],
        ['Data Factory - Orchestration', '$1.00', 'per 1,000 activity runs', 'Azure Data Factory pricing'],
        ['Data Factory - Data Movement', '$0.25', 'per DIU-hour', 'Azure Data Factory pricing'],
        ['Functions Premium - EP1 vCPU', '$126.29', 'per vCPU-month', 'Azure Functions pricing'],
        ['Functions Premium - EP1 Memory', '$8.979', 'per GB-month (3.5 GB)', 'Azure Functions pricing'],
        ['Functions Premium - EP1 Total', '$158.00', 'per instance-month', 'Calculated (1 vCPU + 3.5 GB)'],
        ['Functions Consumption - Execution', '$0.20', 'per million (after 1M free)', 'Azure Functions pricing'],
        ['Functions Consumption - GB-seconds', '$0.000016', 'per GB-s (after 400K free)', 'Azure Functions pricing'],
        ['Event Grid - Basic Operations', '$0.60', 'per million (after 100K free)', 'Azure Event Grid pricing'],
        ['Service Bus - Standard Base', '$0.0135', 'per hour (~$9.86/mo)', 'Azure Service Bus pricing'],
        ['Service Bus - Standard Ops (excess)', '$0.80', 'per million (13-100M)', 'Azure Service Bus pricing'],
        ['Key Vault - Secret Operations', '$0.03', 'per 10K operations', 'Azure Key Vault pricing'],
        ['Log Analytics - Analytics Logs', '$2.30', 'per GB ingested', 'Azure Monitor pricing'],
        ['Log Analytics - Retention (extra)', '$0.10', 'per GB-month', 'Azure Monitor pricing'],
        ['Private Endpoints', '$0.01', 'per hour (~$7.30/mo)', 'Azure Private Link pricing'],
        ['Private Link - Data Processing', '$0.01', 'per GB (0-1 PB tier)', 'Azure Private Link pricing'],
        ['API Management - Standard v2 Base', '$700.00', 'per month', 'Azure API Management pricing'],
        ['API Management - Standard v2 Requests', '$2.50', 'per million (after 50M)', 'Azure API Management pricing'],
        ['API Management - Standard v2 Scale-out', '$500.00', 'per additional unit', 'Azure API Management pricing'],
        ['API Management - Consumption', '$0.042', 'per 10K operations', 'Azure API Management pricing'],
        ['Purview - Data Map CU (placeholder)', '$190.00', 'per CU-month', 'Estimate (needs confirmation)'],
    ]
    
    for row, data in enumerate(pricing_data, start=5):
        for col, value in enumerate(data, start=1):
            ws.cell(row=row, column=col, value=value)
    
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 35

def create_monthly_costs_sheet(wb):
    """Detailed monthly cost breakdown by service"""
    ws = wb.create_sheet("Monthly Costs (Prod)")
    
    ws['A1'] = "Production Environment - Monthly Cost Breakdown"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = "All figures in USD"
    ws['A2'].font = Font(size=10, italic=True)
    
    headers = ['Category', 'Low', 'Expected', 'High', 'Basis / Formula']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    cost_data = [
        ['Storage (Landing + Raw + Outbound)', '$18', '$25', '$45', 'Hot 250 GB * $0.018 + Cool 500 GB * $0.01 + misc'],
        ['Data Factory', '$150', '$215', '$400', 'Orchestration: 30K runs * $0.001; Data movement: 22K copies (2 DIU * 1 min)'],
        ['Functions (Premium plan baseline)', '$158', '$170', '$320', '1 EP1 instance always on for VNet; High = 2 pre-warmed instances'],
        ['Event Grid', '$0', '$0', '$1', '~22K ops < 100K free; High = ~2M ops'],
        ['Service Bus (Standard)', '$10', '$10', '$40', 'Base ~$9.86; High adds excess ops (20M → 7M billable)'],
        ['Key Vault', '$0.50', '$1', '$3', 'Secret ops + occasional cert ops'],
        ['Purview Data Governance', '$0', '$190', '$380', 'Low = deferred; Expected = 1 CU; High = 2 CUs'],
        ['Log Analytics (+ App Insights)', '$69', '$85', '$138', 'Ingestion GB * $2.30 (30 / 37 / 60 GB)'],
        ['Monitoring & Alerts', '$3', '$5', '$10', 'Alert rule queries & notifications'],
        ['Networking (Private Endpoints)', '$30', '$37', '$60', '4-6 endpoints * $0.01/hr plus Private Link data buffer'],
        ['API Management', '$0', '$700', '$750', 'Low = deferred; Expected = Standard v2 baseline; High = overage calls'],
    ]
    
    row = 5
    for data in cost_data:
        for col, value in enumerate(data, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1
    
    # Total row
    total_row = row
    ws[f'A{total_row}'] = "Total (Prod Monthly)"
    ws[f'A{total_row}'].font = Font(bold=True, size=12)
    ws[f'B{total_row}'] = "$439"
    ws[f'C{total_row}'] = "$1,438"
    ws[f'D{total_row}'] = "$2,147"
    ws[f'E{total_row}'] = "Summation (rounded)"
    
    for col in range(1, 6):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.font = Font(bold=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 70

def create_transaction_volumes_sheet(wb):
    """Transaction volume assumptions and calculations"""
    ws = wb.create_sheet("Transaction Volumes")
    
    ws['A1'] = "Transaction Volume Assumptions"
    ws['A1'].font = Font(size=14, bold=True)
    
    headers = ['Volume Driver', 'Current/Year 0', 'Projection (Jan 1)', 'Notes']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    volume_data = [
        ['Active EDI Processes', '5-10', '10-12', '834, 837/835, plus TA1/999 and outbound assembly'],
        ['Weekly Transactions (inbound files)', '~5,000', 'Scale +15-20% YoY', '≈260,000 / year'],
        ['Annual Claims (837)', '~250,000', 'Could rise with payer feeds', 'Drives storage + routing + ack volume'],
        ['Subscribers / Members', '7,200', '10,000', 'Used for 834 enrollment updates'],
        ['X12 Sets Phase 1', '834, 837, 835, TA1, 999, 277CA', 'Add 271, 277, 278 later', 'Outbound responses increase volume'],
        ['Average File Size', '1-5 MB (peaks 50-100 MB)', 'Similar', 'Larger 837/835 batch peaks'],
    ]
    
    for row, data in enumerate(volume_data, start=4):
        for col, value in enumerate(data, start=1):
            ws.cell(row=row, column=col, value=value)
    
    # Monthly breakdown
    ws['A12'] = "Monthly Volume Estimates (Production)"
    ws['A12'].font = Font(size=14, bold=True)
    
    headers2 = ['Metric', 'Volume', 'Calculation']
    for col, header in enumerate(headers2, start=1):
        cell = ws.cell(row=14, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    monthly_data = [
        ['Inbound Files', '~21,700', '5,000/week * 52 / 12'],
        ['ADF Activity Runs', '~28,000', '21,700 * 1.3 (includes lookups & retries)'],
        ['Function Invocations (Router)', '~22,000', '~1 per inbound file'],
        ['Function Invocations (Orchestrator)', '~6,000', '192 batches/day (5-min intervals, 16 hrs)'],
        ['Service Bus Messages (Routing)', '~76,000', 'Avg 3.5 ST sets per file * 21,700'],
        ['Service Bus Operations (Total)', '~400,000', 'Publish + deliveries + management'],
        ['Event Grid Events', '~21,700', 'Blob Created triggers'],
        ['Log Analytics Ingestion', '30-45 GB', '1.0-1.5 GB/day * 30 days'],
        ['Storage Growth (Raw)', '~65 GB', '21,700 files * 3 MB avg'],
        ['API Calls (Expected)', '~5,000', 'Assuming 10% of files arrive via API vs SFTP'],
    ]
    
    for row, data in enumerate(monthly_data, start=15):
        for col, value in enumerate(data, start=1):
            ws.cell(row=row, column=col, value=value)
    
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50

def create_environment_rollup_sheet(wb):
    """Multi-environment cost rollup"""
    ws = wb.create_sheet("Environment Roll-Up")
    
    ws['A1'] = "All Environments - Monthly Cost Roll-Up"
    ws['A1'].font = Font(size=14, bold=True)
    
    headers = ['Environment', 'Low', 'Expected', 'High', 'Notes']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    env_data = [
        ['Dev (~25%)', '$172', '$424', '$611', 'Includes shared Premium Functions + Private Endpoints + APIM allocation'],
        ['Test (~40%)', '$209', '$603', '$877', 'Load/replay tests with shared Premium Functions + Private Endpoints + APIM'],
        ['Prod', '$439', '$1,438', '$2,147', 'From detailed breakdown'],
    ]
    
    for row, data in enumerate(env_data, start=4):
        for col, value in enumerate(data, start=1):
            ws.cell(row=row, column=col, value=value)
    
    # Total row
    ws['A7'] = "Total / Month"
    ws['B7'] = "$820"
    ws['C7'] = "$2,465"
    ws['D7'] = "$3,635"
    ws['E7'] = "All environments combined"
    
    for col in range(1, 6):
        cell = ws.cell(row=7, column=col)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.font = Font(bold=True)
    
    # Annual calculations
    ws['A10'] = "Annual Costs (Expected Scenario)"
    ws['A10'].font = Font(size=14, bold=True)
    
    ws['A12'] = "Component"
    ws['B12'] = "Amount"
    for col in range(1, 3):
        cell = ws.cell(row=12, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    annual_data = [
        ['Monthly Total (Expected)', '$2,465'],
        ['Annual (×12)', '$29,580'],
        ['Contingency (20%)', '$5,916'],
        ['Total Budget Ask', '$35,496'],
    ]
    
    for row, data in enumerate(annual_data, start=13):
        ws[f'A{row}'] = data[0]
        ws[f'B{row}'] = data[1]
        if 'Total Budget' in data[0]:
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'B{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 60

def create_sensitivity_analysis_sheet(wb):
    """Cost sensitivity analysis"""
    ws = wb.create_sheet("Sensitivity Analysis")
    
    ws['A1'] = "Cost Sensitivity Levers"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = "Understanding what drives cost changes"
    ws['A2'].font = Font(size=10, italic=True)
    
    headers = ['Driver', 'Elasticity', 'Impact', 'Comment']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    sensitivity_data = [
        ['ST transactions per file (× factor)', 'High', 'ADF + SB + Functions', 'Each extra ST adds routing message + log events'],
        ['Log verbosity (% increase)', 'Linear', '$2.30 per GB', 'Keep below 2 GB/day to maintain <$140/mo'],
        ['Functions Premium instance count', 'Step', '+$158/mo per instance', 'Review before enabling pre-warmed scaling'],
        ['Private Endpoint footprint', 'Linear', '+$7.30/mo per endpoint', 'Each additional endpoint + data processing'],
        ['API Management call volume', 'Step', '+$2.50 per 1M over 50M', 'Standard v2 includes 50M requests/mo'],
        ['Purview adoption breadth', 'Step', '+$190/mo per CU', 'Auto-scans across many sources can double cost'],
        ['Reprocessing rate (%)', 'Moderate', 'Proportional ADF + Functions', '5%→10% increases activity runs'],
        ['Lifecycle policy delay (days)', 'Low-Moderate', 'Storage tier mix', 'Extends Hot storage segment'],
    ]
    
    for row, data in enumerate(sensitivity_data, start=5):
        for col, value in enumerate(data, start=1):
            ws.cell(row=row, column=col, value=value)
    
    # Scaling triggers
    ws['A15'] = "Scaling Triggers & Thresholds"
    ws['A15'].font = Font(size=14, bold=True)
    
    headers2 = ['Trigger Metric', 'Threshold', 'Action', 'Impact']
    for col, header in enumerate(headers2, start=1):
        cell = ws.cell(row=17, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    trigger_data = [
        ['Routing messages', '> 1M/month sustained', 'Evaluate Premium Service Bus', '+$300-$500/mo'],
        ['Log ingestion', '> 2 GB/day trending up', 'Increase sampling / reduce verbosity', 'Savings 10-30%'],
        ['ADF activity concurrency waits', '> 5% weekly avg', 'Split pipelines or add parallelization', 'Maintain SLA'],
        ['Function cold start p95', '> 2s for >5% invocations', 'Add second EP1 instance', '+$158/mo per instance'],
        ['API Management requests', '> 40M/mo approaching 50M', 'Review throttling or scale-out', '+$500/mo per unit'],
        ['Raw storage', '> 2 TB Year 1', 'Accelerate lifecycle to Cool', 'Save 20-30% storage'],
        ['Purview assets', '> 5k catalog expansion', 'Add 1 more CU or optimize schedule', '+$190/mo'],
    ]
    
    for row, data in enumerate(trigger_data, start=18):
        for col, value in enumerate(data, start=1):
            ws.cell(row=row, column=col, value=value)
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30

if __name__ == "__main__":
    create_budget_spreadsheet()
